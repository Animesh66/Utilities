import openpyxl as xl


def get_row_count(file_path, sheet_name):
    workbook = xl.load_workbook(file_path)
    sheet = workbook[sheet_name]
    row_count = sheet.max_row
    return row_count


def get_column_count(file_path, sheet_name):
    workbook = xl.load_workbook(file_path)
    sheet = workbook[sheet_name]
    column_count = sheet.max_column
    return column_count


def get_cell_data(file_path, sheet_name, row, column):
    workbook = xl.load_workbook(file_path)
    sheet = workbook[sheet_name]
    cell = sheet.cell(row=row, column=column)
    cell_value = cell.value
    return cell_value


def set_cell_data(file_path, sheet_name, row, column, data):
    workbook = xl.load_workbook(file_path)
    sheet = workbook[sheet_name]
    cell = sheet.cell(row=row, column=column)
    cell_value = cell.value
    cell_value = data
    workbook.save(file_path)
